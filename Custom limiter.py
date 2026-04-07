import streamlit as st
import pandas as pd
import io
import time
import plotly.express as px
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- 1. UI CONFIGURATION ---
st.set_page_config(page_title="VMS Universal Reporting", layout="wide")
MASTER_PASSWORD = "VMS@123"

st.markdown("""
    <style>
    .stApp {
        background: linear-gradient(rgba(15, 23, 42, 0.92), rgba(15, 23, 42, 0.92)), 
                    url("https://images.unsplash.com/photo-1451187580459-43490279c0fa?ixlib=rb-1.2.1&auto=format&fit=crop&w=1950&q=80");
        background-size: cover; background-attachment: fixed;
    }
    .welcome-note { 
        background: linear-gradient(to right, #00d2ff, #92fe9d); 
        -webkit-background-clip: text; -webkit-text-fill-color: transparent; 
        font-size: 48px !important; font-weight: 700; text-align: center; margin: 40px 0 10px 0;
    }
    .glass-metric {
        background: rgba(255, 255, 255, 0.03); backdrop-filter: blur(15px);
        border: 1px solid rgba(255, 255, 255, 0.1); border-radius: 15px;
        padding: 25px; margin: 10px 0; text-align: center;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
    }
    .metric-value { font-size: 42px; font-weight: 800; color: #92fe9d; }
    </style>
    """, unsafe_allow_html=True)

# --- 2. AUTHENTICATION ---
if 'authenticated' not in st.session_state: st.session_state.authenticated = False
if not st.session_state.authenticated:
    st.markdown('<p class="welcome-note">VMS Reporting System</p>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1.4, 1])
    with col2:
        p = st.text_input("Password", type="password")
        if st.button("Access Dashboard", use_container_width=True):
            if p == MASTER_PASSWORD: st.session_state.authenticated = True; st.rerun()
            else: st.error("Access Denied")
    st.stop()

# --- 3. CORE LOGIC ---
KEYWORDS_TO_IGNORE = ["BADMINTON", "BASKETBALL", "CROSS FITNESS", "SWIMMING", "ZUMBA", "TABLE TENNIS", 
                      "FREESLOT", "FREE SLOT", "SOFT SKILL", "ATOM", "DSA"]
ATT_COL_NAME = "Attended Hours with Approved Leave Percentage"

def is_valid_subject(subject_name):
    s_upper = str(subject_name).upper()
    return not any(bad in s_upper for bad in KEYWORDS_TO_IGNORE)

def get_bracket_summary(data_df, cols, subjects, thresh_to):
    summary_data = []
    for sub in subjects:
        sub_vals = pd.to_numeric(data_df[data_df[cols['subject']] == sub][cols['attendance']], errors='coerce').dropna()
        
        b1 = len(sub_vals[(sub_vals >= 0) & (sub_vals < 50)])
        b2 = len(sub_vals[(sub_vals >= 50) & (sub_vals < 60)])
        b3 = len(sub_vals[(sub_vals >= 60) & (sub_vals < 70)])
        b4 = len(sub_vals[(sub_vals >= 70) & (sub_vals < 75)])
        
        row = {"Subject": sub}
        total = 0
        
        if thresh_to > 0:
            row["0.00-49.99"] = b1
            total += b1
        if thresh_to > 50:
            row["50.00-59.99"] = b2
            total += b2
        if thresh_to > 60:
            row["60.00-69.99"] = b3
            total += b3
        if thresh_to > 70:
            row["70.00-74.99"] = b4
            total += b4
            
        row["Total"] = total
        summary_data.append(row)
        
    return pd.DataFrame(summary_data)

def apply_styles(ws, thresh_to, is_summary=False):
    thin = Side(style='thin', color="4D4D4D")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    crit_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid") 
    warn_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") 
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font, cell.fill, cell.border = Font(bold=True, color="FFFFFF"), h_fill, border
        ws.column_dimensions[cell.column_letter].width = 20

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border, cell.alignment = border, Alignment(horizontal="center")
            if ws.cell(row=1, column=cell.column).value == "Total":
                cell.font = Font(bold=True)
            
            if not is_summary and cell.column > 5:
                try:
                    val = float(cell.value)
                    if val < 70: cell.fill, cell.font = crit_fill, Font(bold=True, color="FFFFFF")
                    elif 70 <= val < thresh_to: cell.fill, cell.font = warn_fill, Font(bold=True, color="000000")
                except: pass

def process_grid(data_df, cols, batch_subjects, low_thresh, high_thresh, show_all=False):
    if data_df.empty: return None, None
    data_df = data_df.copy()
    data_df[cols['attendance']] = pd.to_numeric(data_df[cols['attendance']], errors='coerce')
    
    full_grid = data_df.pivot_table(index=[cols['roll'], cols['name'], cols['batch'], cols['sem']],
                                    columns=cols['subject'], values=cols['attendance'], sort=False).reset_index()
    
    final_subjects = [s for s in batch_subjects if is_valid_subject(s)]
    
    for sub in final_subjects:
        if sub not in full_grid.columns: full_grid[sub] = None
        full_grid[sub] = pd.to_numeric(full_grid[sub], errors='coerce')

    theory_cols = [c for c in final_subjects if not any(x in str(c).upper() for x in ["LAB", "PRACTICAL", "WORKSHOP"])]
    full_grid['Theory Avg'] = full_grid[theory_cols].mean(axis=1).round(2)
    full_grid['Final Avg'] = full_grid[final_subjects].mean(axis=1).round(2)
    
    if show_all:
        shortage_grid = full_grid.copy()
    else:
        # Strict Range Filtering Logic
        mask = (full_grid[final_subjects] >= low_thresh) & (full_grid[final_subjects] <= high_thresh)
        shortage_grid = full_grid[mask.any(axis=1)].copy()
    
    if shortage_grid.empty: return None, None
    
    # Calculate counts strictly within range
    shortage_grid['Subjects in Range'] = ((shortage_grid[final_subjects] >= low_thresh) & (shortage_grid[final_subjects] <= high_thresh)).sum(axis=1)
    sub_counts = ((shortage_grid[final_subjects] >= low_thresh) & (shortage_grid[final_subjects] <= high_thresh)).sum()
    
    if not show_all:
        for sub in final_subjects:
            # Mask out values that are NOT in the chosen range
            shortage_grid[sub] = shortage_grid[sub].apply(lambda x: x if (pd.notnull(x) and low_thresh <= x <= high_thresh) else "")
    
    shortage_grid.insert(0, 'Sl No.', range(1, len(shortage_grid) + 1))
    final_cols = ['Sl No.', cols['roll'], cols['name'], cols['batch'], cols['sem']] + final_subjects + ['Subjects in Range', 'Theory Avg', 'Final Avg']
    
    # Footer row logic
    count_row = pd.DataFrame([["", "", "", "", f"Count ({low_thresh}%-{high_thresh}%)"] + [sub_counts[s] for s in final_subjects] + ["", "", ""]], columns=final_cols)
    shortage_grid = pd.concat([shortage_grid, count_row], ignore_index=True)
    
    return shortage_grid, sub_counts

# --- 4. DASHBOARD INTERFACE ---
uploaded_file = st.file_uploader("📂 Upload Universal Attendance File", type=["xlsx"])

if uploaded_file:
    df_preview = pd.read_excel(uploaded_file, header=None).head(15)
    h_row = 0
    for i, row in df_preview.iterrows():
        if any("ROLL NO" in str(x).upper() for x in row.values):
            h_row = i
            break
    
    df = pd.read_excel(uploaded_file, header=h_row)
    c_map = {'sem': df.columns[5]} 
    for c in df.columns:
        cs = str(c).strip()
        if "Roll No" in cs: c_map['roll'] = c
        elif "Student Name" in cs: c_map['name'] = c
        elif "Batch" in cs: c_map['batch'] = c
        elif any(x in cs for x in ["Course", "Subject"]): c_map['subject'] = c
        elif ATT_COL_NAME in cs: c_map['attendance'] = c

    df = df[df[c_map['subject']].apply(is_valid_subject)]
    df['Dept'] = df[c_map['batch']].astype(str).apply(lambda x: x.split()[0].upper())
    all_subjects = sorted(df[c_map['subject']].unique())
    
    with st.sidebar:
        st.markdown("### 🛠️ Global Parameters")
        
        # Two distinct buttons for Range Input
        col_f, col_t = st.columns(2)
        with col_f:
            low_val = st.number_input("From (%)", 0.00, 100.00, 0.00, 0.01, format="%.2f")
        with col_t:
            high_val = st.number_input("To (%)", 0.00, 100.00, 75.00, 0.01, format="%.2f")
        
        dept_choice = st.selectbox("Select Department", ["All Departments"] + sorted(df['Dept'].unique()))
        
        st.divider()
        st.markdown("### 🔍 Exclusion Filters")
        exclude_subjects = st.multiselect("Exclude Subjects/Faculty", all_subjects)
        
        if st.button("Logout"): st.session_state.authenticated = False; st.rerun()

    if exclude_subjects:
        df = df[~df[c_map['subject']].isin(exclude_subjects)]
        
    if dept_choice != "All Departments":
        df = df[df['Dept'] == dept_choice]
        active_depts = [dept_choice]
    else:
        active_depts = sorted(df['Dept'].unique())

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Audit Log
        pd.DataFrame({"Filter Criteria": [f"Dept: {dept_choice}", f"Range: {low_val}% to {high_val}%"]}).to_excel(writer, sheet_name="Audit", index=False)
        summaries, subject_impact = [], pd.Series(dtype=float)
        
        tabs = st.tabs(["📊 COMMAND CENTER"] + [f"💎 {d}" for d in active_depts])

        for d_idx, dept in enumerate(active_depts):
            d_df = df[df['Dept'] == dept]
            unique_batches = d_df[c_map['batch']].astype(str).unique()
            series_list = set()
            for b in unique_batches:
                b_upper = b.upper()
                b_parts = b.split()
                if "BCA" in b_upper:
                    year = next((p for p in b_parts if p.isdigit()), "Series")
                    series_list.add(f"BCA {year}")
                elif "MCA" in b_upper:
                    year = next((p for p in b_parts if p.isdigit()), "Series")
                    series_list.add(f"MCA {year}")
                else:
                    series_list.add(' '.join(b_parts[:2]))
            
            series_list = sorted(list(series_list))
            
            with tabs[d_idx+1]:
                for series in series_list:
                    s_df = d_df[d_df[c_map['batch']].astype(str).str.contains(series.split()[0]) & 
                                d_df[c_map['batch']].astype(str).str.contains(series.split()[-1])]
                    
                    s_subs = sorted([s for s in s_df[c_map['subject']].unique() if is_valid_subject(s)])
                    
                    # 1. RANGE SUMMARY SHEET (Filtered)
                    gen_grid, _ = process_grid(s_df, c_map, s_subs, low_val, high_val, show_all=False)
                    if gen_grid is not None:
                        with st.expander(f"👁️ {series} ({low_val}% - {high_val}%)"):
                            st.dataframe(gen_grid, hide_index=True, use_container_width=True)
                        sn = f"{series} Range"[:31]
                        gen_grid.to_excel(writer, sheet_name=sn, index=False)
                        get_bracket_summary(s_df, c_map, s_subs, high_val).to_excel(writer, sheet_name=sn, startrow=len(gen_grid)+2, index=False)
                        apply_styles(writer.sheets[sn], high_val)

                    # 2. FULL GRID SHEET (All Data for context)
                    all_grid, _ = process_grid(s_df, c_map, s_subs, low_val, high_val, show_all=True)
                    if all_grid is not None:
                        sn_all = f"{series} FULL"[:31]
                        all_grid.to_excel(writer, sheet_name=sn_all, index=False)
                        apply_styles(writer.sheets[sn_all], high_val)
                    
                    # 3. INDIVIDUAL SECTION SHEETS
                    sections = sorted(s_df[c_map['batch']].unique())
                    for sec in sections:
                        sec_df = s_df[s_df[c_map['batch']] == sec]
                        grid, counts = process_grid(sec_df, c_map, s_subs, low_val, high_val)
                        if grid is not None:
                            sn_sec = str(sec).replace("/", "-")[:31]
                            grid.to_excel(writer, sheet_name=sn_sec, index=False)
                            apply_styles(writer.sheets[sn_sec], high_val)
                            summaries.append({'Section': sec, 'Count': len(grid)-1})
                            subject_impact = subject_impact.add(counts, fill_value=0)

        # Dashboard Summary
        with tabs[0]:
            if summaries:
                sum_df = pd.DataFrame(summaries)
                m_cols = st.columns(min(len(sum_df), 4))
                for idx, row in sum_df.iterrows():
                    with m_cols[idx % 4]:
                        st.markdown(f'<div class="glass-metric"><div class="metric-title">{row["Section"]}</div><div class="metric-value">{row["Count"]}</div></div>', unsafe_allow_html=True)
                
                c1, c2 = st.columns(2)
                with c1: st.plotly_chart(px.bar(sum_df, x='Section', y='Count', title="Students in Range per Section", template="plotly_dark"), use_container_width=True)
                with c2:
                    if not subject_impact.empty and subject_impact.sum() > 0:
                        impact_df = subject_impact.reset_index()
                        impact_df.columns = ['Subject', 'Students']
                        st.plotly_chart(px.pie(impact_df.head(10), names='Subject', values='Students', hole=0.4, title="Subject Impact (In Range)", template="plotly_dark"), use_container_width=True)
                sum_df.to_excel(writer, sheet_name='SUMMARY', index=False)
            else:
                st.info(f"No students found within the range of {low_val}% to {high_val}%.")

    st.download_button(f"📥 Download {dept_choice} Report", output.getvalue(), f"VMS_{dept_choice}_Range_Report.xlsx", use_container_width=True)
